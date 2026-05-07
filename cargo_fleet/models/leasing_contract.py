from dateutil.relativedelta import relativedelta

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
import base64
import io
from datetime import datetime, date
from openpyxl import load_workbook


class LeasingEcheance(models.Model):
    _name = 'leasing.echeance'
    _description = 'Echéance Leasing'
    _order = 'sequence asc, date_echeance asc, id asc'

    purchase_order_id = fields.Many2one(
        'purchase.order',
        string="Contrat leasing",
        required=True,
        ondelete='cascade'
    )

    vehicle_id = fields.Many2one(
        'fleet.vehicle',
        string="Véhicule",
        related='purchase_order_id.vehicle_id',
        store=True,
        readonly=True
    )

    license_plate = fields.Char(
        string="Immatriculation",
        related='vehicle_id.license_plate',
        store=True,
        readonly=True
    )

    numero_w = fields.Char(
        string="Numéro W",
        related='vehicle_id.numero_w',
        store=True,
        readonly=True
    )

    model_name = fields.Char(
        string="Modèle",
        related='vehicle_id.model_name',
        store=True,
        readonly=True
    )

    brand_name = fields.Char(
        string="Marque",
        related='vehicle_id.brand_name',
        store=True,
        readonly=True
    )

    partner_id = fields.Many2one(
        'res.partner',
        string="Société leasing",
        related='purchase_order_id.leasing_company_id',
        store=True,
        readonly=True
    )

    leasing_contract_status = fields.Selection(
        related='purchase_order_id.leasing_contract_status',
        string="Statut contrat",
        store=True,
        readonly=True
    )

    sequence = fields.Integer(string="N° échéance", required=True, default=1)
    name = fields.Char(string="Libellé", compute="_compute_name", store=True)

    date_echeance = fields.Date(string="Date échéance", required=True)

    date_prelevement_reel = fields.Date(
        string="Date prélèvement réel",
        copy=False
    )

    payment_mode = fields.Selection([
        ('manuel', 'Manuel'),
        ('prelevement', 'Prélèvement'),
    ], string="Mode paiement", default='prelevement')

    capital_amount = fields.Float(string="Capital", default=0.0)
    interest_amount = fields.Float(string="Intérêts", default=0.0)

    amount_total = fields.Float(
        string="Montant prélèvement TTC",
        default=0.0
    )

    tva_import_amount = fields.Float(
        string="TVA Excel",
        default=0.0
    )

    amount_penalite = fields.Float(
        string="Montant pénalité",
        default=0.0
    )

    amount_ht = fields.Float(
        string="Montant prélèvement HT",
        compute="_compute_amounts_from_taxes",
        store=True
    )

    tax_ids = fields.Many2many(
        'account.tax',
        'leasing_echeance_account_tax_rel',
        'echeance_id',
        'tax_id',
        string="Taxes",
        domain="[('type_tax_use', 'in', ['purchase', 'none'])]"
    )

    tva_amount = fields.Float(
        string="Montant TVA",
        compute="_compute_amounts_from_taxes",
        store=True
    )

    payment_ids = fields.One2many(
        'leasing.payment',
        'echeance_id',
        string="Paiements"
    )

    amount_paid = fields.Float(
        string="Montant payé",
        compute="_compute_payment_amounts",
        store=True
    )

    amount_due = fields.Float(
        string="Reste à payer",
        compute="_compute_payment_amounts",
        store=True
    )

    payment_date = fields.Date(
        string="Date dernier paiement",
        compute="_compute_payment_date",
        store=True
    )

    state = fields.Selection([
        ('draft', 'Non payé'),
        ('partial', 'Partiellement payé'),
        ('paid', 'Payé'),
        ('late', 'En retard'),
    ], string="Statut", compute="_compute_state", store=True)

    note = fields.Text(string="Observation")

    bill_purchase_order_id = fields.Many2one(
        'purchase.order',
        string="BL échéance",
        readonly=True,
        copy=False
    )

    vendor_bill_id = fields.Many2one(
        'account.move',
        string="Facture échéance",
        readonly=True,
        copy=False
    )

    penalty_bill_purchase_order_id = fields.Many2one(
        'purchase.order',
        string="BL pénalité",
        readonly=True,
        copy=False
    )

    penalty_vendor_bill_id = fields.Many2one(
        'account.move',
        string="Facture pénalité",
        readonly=True,
        copy=False
    )

    can_create_bill_order = fields.Boolean(
        string="Peut créer BL",
        compute="_compute_can_create_bill_order"
    )

    def action_confirm_and_pay_selected(self):
        if not self:
            raise UserError(_("Veuillez sélectionner au moins une échéance."))

        invalides = self.filtered(lambda r: r.state not in ('draft', 'partial', 'late'))
        if invalides:
            raise UserError(_("Certaines échéances sélectionnées ne sont pas payables."))

        for echeance in self:
            echeance.action_confirm_and_pay()

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

    @api.depends('sequence', 'vehicle_id')
    def _compute_name(self):
        for rec in self:
            vehicle = rec.vehicle_id.display_name if rec.vehicle_id else ''
            rec.name = f"Echéance {rec.sequence}" + (f" - {vehicle}" if vehicle else "")

    @api.onchange('date_echeance')
    def _onchange_date_echeance_set_real_debit_date(self):
        for rec in self:
            if rec.date_echeance and not rec.date_prelevement_reel:
                rec.date_prelevement_reel = rec.date_echeance

    @api.depends('amount_total', 'tva_import_amount')
    def _compute_amounts_from_taxes(self):
        for rec in self:
            ttc = rec.amount_total or 0.0
            tva = rec.tva_import_amount or 0.0

            rec.tva_amount = tva
            rec.amount_ht = ttc - tva

    @api.depends('payment_ids.amount', 'amount_total')
    def _compute_payment_amounts(self):
        for rec in self:
            paid = sum(rec.payment_ids.mapped('amount'))
            rec.amount_paid = paid
            rec.amount_due = max((rec.amount_total or 0.0) - paid, 0.0)

    @api.depends('payment_ids.payment_date')
    def _compute_payment_date(self):
        for rec in self:
            dates = rec.payment_ids.mapped('payment_date')
            rec.payment_date = max(dates) if dates else False

    @api.depends('amount_due', 'amount_paid', 'amount_total', 'date_echeance')
    def _compute_state(self):
        today = fields.Date.context_today(self)
        for rec in self:
            if rec.amount_total and rec.amount_due <= 0:
                rec.state = 'paid'
            elif rec.amount_paid > 0 and rec.amount_due > 0:
                rec.state = 'partial'
            elif rec.date_echeance and rec.date_echeance < today:
                rec.state = 'late'
            else:
                rec.state = 'draft'

    @api.depends('bill_purchase_order_id', 'state', 'amount_due')
    def _compute_can_create_bill_order(self):
        for rec in self:
            rec.can_create_bill_order = bool(
                rec.purchase_order_id
                and rec.purchase_order_id.vehicle_purchase_type == 'leasing_contract'
                and not rec.purchase_order_id.is_leasing_bill_order
                and not rec.bill_purchase_order_id
                and rec.state != 'paid'
                and rec.amount_due > 0
            )

    @api.constrains(
        'amount_total',
        'capital_amount',
        'interest_amount',
        'tva_amount',
        'amount_ht',
        'amount_penalite',
        'tva_import_amount'
    )
    def _check_positive_amounts(self):
        for rec in self:
            if (
                rec.capital_amount < 0
                or rec.interest_amount < 0
                or rec.tva_amount < 0
                or rec.amount_total < 0
                or rec.amount_ht < 0
                or rec.amount_penalite < 0
                or rec.tva_import_amount < 0
            ):
                raise ValidationError(_("Les montants d'une échéance ne peuvent pas être négatifs."))

            if rec.tva_import_amount > rec.amount_total:
                raise ValidationError(_("La TVA ne peut pas dépasser le montant TTC."))

    @api.constrains('date_echeance', 'date_prelevement_reel')
    def _check_date_prelevement_reel(self):
        for rec in self:
            if not rec.date_echeance or not rec.date_prelevement_reel:
                continue

            min_date = rec.date_echeance - relativedelta(days=10)
            max_date = rec.date_echeance + relativedelta(days=60)

            if rec.date_prelevement_reel < min_date or rec.date_prelevement_reel > max_date:
                raise ValidationError(_(
                    "La date de prélèvement réel doit être comprise entre 10 jours avant "
                    "et 60 jours après la date d'échéance."
                ))

    def _get_effective_payment_date(self):
        self.ensure_one()
        return self.date_prelevement_reel or self.date_echeance or fields.Date.context_today(self)

    def _ensure_bill_purchase_order(self, is_penalty=False):
        self.ensure_one()

        if is_penalty and self.penalty_bill_purchase_order_id:
            return self.penalty_bill_purchase_order_id

        if not is_penalty and self.bill_purchase_order_id:
            return self.bill_purchase_order_id

        if not self.purchase_order_id or self.purchase_order_id.vehicle_purchase_type != 'leasing_contract':
            raise ValidationError(_("Cette action est réservée aux échéances d’un contrat leasing."))

        if self.purchase_order_id.is_leasing_bill_order:
            raise ValidationError(_("Impossible de créer un BL depuis un BL d'échéance."))

        if not is_penalty and self.amount_due <= 0:
            raise ValidationError(_("Cette échéance est déjà soldée."))

        if is_penalty and (self.amount_penalite or 0.0) <= 0:
            raise ValidationError(_("Le montant de pénalité doit être strictement positif."))

        contract = self.purchase_order_id

        if not contract.vehicle_id:
            raise ValidationError(_("Aucun véhicule n’est lié au contrat leasing."))

        if is_penalty:
            line_vals = contract._prepare_leasing_penalty_order_line_vals(self)
            origin_label = "Pénalité"
        else:
            line_vals = contract._prepare_leasing_bill_order_line_vals(self)
            origin_label = "Echéance"

        po_vals = {
            'partner_id': contract.leasing_company_id.id or contract.partner_id.id,
            'vehicle_id': contract.vehicle_id.id,
            'vehicle_purchase_type': 'leasing_contract',
            'is_leasing_bill_order': True,
            'source_leasing_contract_id': contract.id,
            'source_leasing_echeance_id': self.id,
            'date_order': fields.Datetime.now(),
            'origin': "%s - %s %s - %s" % (
                contract.name or contract.leasing_contract_number or "Contrat leasing",
                origin_label,
                self.sequence,
                self.date_echeance or ''
            ),
            'order_line': [(0, 0, line_vals)],
        }

        bill_po = self.env['purchase.order'].create(po_vals)

        if is_penalty:
            self.penalty_bill_purchase_order_id = bill_po.id
        else:
            self.bill_purchase_order_id = bill_po.id

        return bill_po

    def _ensure_confirmed_bill_purchase_order(self, is_penalty=False):
        self.ensure_one()
        po = self._ensure_bill_purchase_order(is_penalty=is_penalty)
        if po.state not in ('purchase', 'done'):
            po.button_confirm()
        return po

    def _find_existing_vendor_bill(self, po):
        self.ensure_one()
        return self.env['account.move'].search([
            ('move_type', '=', 'in_invoice'),
            ('invoice_line_ids.purchase_line_id.order_id', '=', po.id),
        ], order='id desc', limit=1)

    def _find_or_create_vendor_bill(self, is_penalty=False):
        self.ensure_one()

        if is_penalty and self.penalty_vendor_bill_id:
            return self.penalty_vendor_bill_id

        if not is_penalty and self.vendor_bill_id:
            return self.vendor_bill_id

        po = self._ensure_confirmed_bill_purchase_order(is_penalty=is_penalty)

        existing_bill = self._find_existing_vendor_bill(po)
        if existing_bill:
            if is_penalty:
                self.penalty_vendor_bill_id = existing_bill.id
            else:
                self.vendor_bill_id = existing_bill.id
            return existing_bill

        if hasattr(po, 'action_create_vendor_bill'):
            po.action_create_vendor_bill()
        else:
            raise ValidationError(_("Impossible de créer automatiquement la facture fournisseur depuis ce bon de commande."))

        bill = self._find_existing_vendor_bill(po)
        if not bill:
            raise ValidationError(_("Aucune facture fournisseur n’a été générée."))

        if is_penalty:
            self.penalty_vendor_bill_id = bill.id
        else:
            self.vendor_bill_id = bill.id

        return bill

    def _post_vendor_bill_if_needed(self, bill):
        self.ensure_one()

        effective_date = self._get_effective_payment_date()

        if bill.state == 'draft':
            if not bill.invoice_date:
                bill.invoice_date = effective_date
            if not bill.date:
                bill.date = effective_date
            bill.action_post()

        return bill

    def _pay_bill(self, bill):
        self.ensure_one()

        contract = self.purchase_order_id

        if not contract.payment_journal_id:
            raise ValidationError(_("Veuillez choisir la banque de paiement sur le contrat."))

        if not contract.contract_bank_account_id:
            raise ValidationError(_("Veuillez choisir le numéro de compte bénéficiaire sur le contrat."))

        payment_register = self.env['account.payment.register'].with_context(
            active_model='account.move',
            active_ids=bill.ids
        ).create({
            'payment_date': self._get_effective_payment_date(),
            'journal_id': contract.payment_journal_id.id,
        })

        payment_register.action_create_payments()

    def action_confirm_and_pay(self):
        tracking = self.env['fleet.location.tracking']

        for rec in self:
            contract = rec.purchase_order_id

            if not contract.payment_journal_id:
                raise ValidationError(_("Veuillez choisir la banque de paiement sur le contrat avant de lancer le paiement automatique."))

            if not contract.contract_bank_account_id:
                raise ValidationError(_("Veuillez choisir le numéro de compte bénéficiaire sur le contrat avant de lancer le paiement automatique."))

            effective_date = rec._get_effective_payment_date()

            normal_po = rec._ensure_confirmed_bill_purchase_order(is_penalty=False)
            normal_bill = rec._find_or_create_vendor_bill(is_penalty=False)
            normal_bill = rec._post_vendor_bill_if_needed(normal_bill)
            rec._pay_bill(normal_bill)

            tracking.sync_from_leasing_echeance(rec, normal_bill)

            existing_normal_payment = rec.payment_ids.filtered(
                lambda p: p.payment_date == effective_date
                and p.amount == rec.amount_total
                and p.reference == (normal_bill.name or normal_bill.ref or normal_po.name)
            )[:1]

            if not existing_normal_payment:
                self.env['leasing.payment'].create({
                    'echeance_id': rec.id,
                    'payment_date': effective_date,
                    'amount': rec.amount_total,
                    'payment_method': 'prelevement',
                    'reference': normal_bill.name or normal_bill.ref or normal_po.name,
                    'note': _("Paiement automatique échéance leasing"),
                })

            if (rec.amount_penalite or 0.0) > 0:
                penalty_po = rec._create_penalty_expense_if_needed()

                penalty_bill = penalty_po._get_related_vendor_bills().filtered(
                    lambda m: m.move_type == 'in_invoice'
                )[:1]

                if not penalty_bill:
                    penalty_po.with_context(skip_fleet_tracking=True).action_create_invoice()
                    penalty_bill = penalty_po._get_related_vendor_bills().filtered(
                        lambda m: m.move_type == 'in_invoice'
                    )[:1]

                if not penalty_bill:
                    raise ValidationError(_("Aucune facture fournisseur de pénalité n’a été générée."))

                rec.penalty_bill_purchase_order_id = penalty_po.id
                rec.penalty_vendor_bill_id = penalty_bill.id

                penalty_bill = rec._post_vendor_bill_if_needed(penalty_bill)
                rec._pay_bill(penalty_bill)

            rec.purchase_order_id._update_leasing_contract_status_from_echeances()

        return True

    def _create_penalty_expense_if_needed(self):
        self.ensure_one()

        if (self.amount_penalite or 0.0) <= 0:
            return False

        contract = self.purchase_order_id

        partner = contract.leasing_company_id or contract.partner_id
        if not partner:
            raise ValidationError(_("Veuillez renseigner la société de leasing / fournisseur."))

        existing_expense = self.env['purchase.order'].search([
            ('vehicle_purchase_type', '=', 'expense'),
            ('fleet_purchase_service_code', '=', 'penalite'),
            ('vehicle_id', '=', contract.vehicle_id.id),
            ('origin', '=', "%s - Pénalité échéance %s" % (
                contract.name or contract.leasing_contract_number or "Contrat leasing",
                self.sequence
            )),
        ], limit=1)

        if existing_expense:
            return existing_expense

        product = self.env['product.product'].search([
            ('purchase_ok', '=', True),
            ('product_tmpl_id.fleet_service_type_id.service_code', '=', 'penalite')
        ], limit=1)

        if not product:
            raise ValidationError(_("Aucun produit achat n'est configuré pour le service 'penalite'."))

        expense_po = self.env['purchase.order'].create({
            'partner_id': partner.id,
            'vehicle_id': contract.vehicle_id.id,
            'vehicle_purchase_type': 'expense',
            'fleet_purchase_service_code': 'penalite',
            'origin': "%s - Pénalité échéance %s" % (
                contract.name or contract.leasing_contract_number or "Contrat leasing",
                self.sequence
            ),
            'order_line': [(0, 0, {
                'product_id': product.id,
                'name': _("Pénalité leasing - Echéance %s") % self.sequence,
                'product_qty': 1.0,
                'product_uom_id': product.uom_id.id,
                'price_unit': self.amount_penalite,
                'date_planned': fields.Datetime.now(),
            })],
        })

        self.env['purchase.order.service.vehicle'].create({
            'purchase_order_id': expense_po.id,
            'service_code': 'penalite',
            'vehicle_id': contract.vehicle_id.id,
            'amount': self.amount_penalite,
            'note': _("Créée automatiquement depuis l'échéance leasing %s") % self.sequence,
        })

        expense_po.button_confirm()
        return expense_po

    def action_create_bill_purchase_order(self):
        self.ensure_one()

        po = self._ensure_bill_purchase_order(is_penalty=False)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Bon de commande'),
            'res_model': 'purchase.order',
            'res_id': po.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_bill_purchase_order(self):
        self.ensure_one()

        if not self.bill_purchase_order_id:
            raise ValidationError(_("Aucun bon de commande n’a encore été créé pour cette échéance."))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Bon de commande'),
            'res_model': 'purchase.order',
            'res_id': self.bill_purchase_order_id.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_vendor_bill(self):
        self.ensure_one()

        if not self.vendor_bill_id:
            raise ValidationError(_("Aucune facture fournisseur n’a encore été créée pour cette échéance."))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Facture fournisseur'),
            'res_model': 'account.move',
            'res_id': self.vendor_bill_id.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_penalty_bill_purchase_order(self):
        self.ensure_one()

        if not self.penalty_bill_purchase_order_id:
            raise ValidationError(_("Aucun bon de commande de pénalité n’a encore été créé."))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Bon de commande pénalité'),
            'res_model': 'purchase.order',
            'res_id': self.penalty_bill_purchase_order_id.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_penalty_vendor_bill(self):
        self.ensure_one()

        if not self.penalty_vendor_bill_id:
            raise ValidationError(_("Aucune facture de pénalité n’a encore été créée."))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Facture fournisseur pénalité'),
            'res_model': 'account.move',
            'res_id': self.penalty_vendor_bill_id.id,
            'view_mode': 'form',
            'target': 'current',
        }


class LeasingPayment(models.Model):
    _name = 'leasing.payment'
    _description = 'Paiement Leasing'
    _order = 'payment_date desc, id desc'

    echeance_id = fields.Many2one(
        'leasing.echeance',
        string="Echéance",
        required=True,
        ondelete='cascade'
    )

    purchase_order_id = fields.Many2one(
        'purchase.order',
        string="Contrat leasing",
        related='echeance_id.purchase_order_id',
        store=True,
        readonly=True
    )

    vehicle_id = fields.Many2one(
        'fleet.vehicle',
        string="Véhicule",
        related='echeance_id.vehicle_id',
        store=True,
        readonly=True
    )

    license_plate = fields.Char(
        string="Immatriculation",
        related='vehicle_id.license_plate',
        store=True,
        readonly=True
    )

    numero_w = fields.Char(
        string="Numéro W",
        related='vehicle_id.numero_w',
        store=True,
        readonly=True
    )

    model_name = fields.Char(
        string="Modèle",
        related='vehicle_id.model_name',
        store=True,
        readonly=True
    )

    brand_name = fields.Char(
        string="Marque",
        related='vehicle_id.brand_name',
        store=True,
        readonly=True
    )

    payment_date = fields.Date(
        string="Date paiement",
        required=True,
        default=fields.Date.context_today
    )

    amount = fields.Float(string="Montant payé", required=True)

    payment_method = fields.Selection([
        ('virement', 'Virement'),
        ('cheque', 'Chèque'),
        ('prelevement', 'Prélèvement'),
        ('especes', 'Espèces'),
        ('autre', 'Autre'),
    ], string="Mode de paiement", default='prelevement')

    reference = fields.Char(string="Référence")
    note = fields.Text(string="Observation")

    attachment = fields.Binary(string="Justificatif")
    attachment_filename = fields.Char(string="Nom justificatif")

    @api.constrains('amount')
    def _check_amount(self):
        for rec in self:
            if rec.amount <= 0:
                raise ValidationError(_("Le montant payé doit être strictement positif."))


class PurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    leasing_tax_ids = fields.Many2many(
        'account.tax',
        'purchase_order_leasing_tax_rel',
        'order_id',
        'tax_id',
        string="Taxes leasing",
        domain="[('type_tax_use', 'in', ['purchase', 'none'])]"
    )

    leasing_amount_debit_ttc = fields.Float(
        string="Montant prélèvement TTC"
    )

    leasing_amount_debit_ht = fields.Float(
        string="Montant prélèvement HT",
        compute="_compute_leasing_debit_amounts",
        store=True
    )

    leasing_debit_tax_amount = fields.Float(
        string="Montant TVA prélèvement",
        compute="_compute_leasing_debit_amounts",
        store=True
    )

    leasing_echeance_ids = fields.One2many(
        'leasing.echeance',
        'purchase_order_id',
        string="Echéancier leasing"
    )

    leasing_payment_ids = fields.One2many(
        'leasing.payment',
        'purchase_order_id',
        string="Paiements leasing"
    )

    leasing_echeance_count = fields.Integer(
        string="Nombre échéances",
        compute="_compute_leasing_summary"
    )

    leasing_paid_echeance_count = fields.Integer(
        string="Echéances payées",
        compute="_compute_leasing_summary"
    )

    leasing_remaining_echeance_count = fields.Integer(
        string="Echéances restantes",
        compute="_compute_leasing_summary"
    )

    leasing_late_echeance_count = fields.Integer(
        string="Echéances en retard",
        compute="_compute_leasing_summary"
    )

    leasing_total_amount = fields.Float(
        string="Total échéancier",
        compute="_compute_leasing_summary"
    )

    leasing_total_paid = fields.Float(
        string="Total payé",
        compute="_compute_leasing_summary"
    )

    leasing_total_remaining = fields.Float(
        string="Reste à payer",
        compute="_compute_leasing_summary"
    )

    leasing_vehicle_return_date = fields.Date(string="Date restitution")
    leasing_allowed_km = fields.Float(string="Kilométrage contractuel")
    leasing_actual_km = fields.Float(string="Kilométrage réel retour")

    leasing_extra_km = fields.Float(
        string="Dépassement KM",
        compute="_compute_km_info",
        store=True
    )

    leasing_extra_km_penalty = fields.Float(string="Pénalité dépassement KM")

    leasing_option_purchase = fields.Boolean(string="Option d'achat levée")
    leasing_option_purchase_date = fields.Date(string="Date levée option")
    leasing_option_purchase_amount = fields.Float(string="Montant levée option")

    leasing_contract_attachment = fields.Binary(string="Contrat signé")
    leasing_contract_attachment_filename = fields.Char(string="Nom contrat signé")

    leasing_amortization_attachment = fields.Binary(string="Tableau amortissement")
    leasing_amortization_attachment_filename = fields.Char(string="Nom tableau amortissement")

    @api.depends('leasing_amount_debit_ttc', 'leasing_tax_ids', 'leasing_tax_ids.amount')
    def _compute_leasing_debit_amounts(self):
        for rec in self:
            ttc = rec.leasing_amount_debit_ttc or 0.0
            total_tax_percent = sum(rec.leasing_tax_ids.mapped('amount'))

            if total_tax_percent:
                ht = ttc / (1 + (total_tax_percent / 100.0))
                tax_amount = ttc - ht
            else:
                ht = ttc
                tax_amount = 0.0

            rec.leasing_amount_debit_ht = ht
            rec.leasing_debit_tax_amount = tax_amount

    def _convert_excel_date(self, value, row_index):
        if not value:
            raise ValidationError(_("Ligne %s : la date d'échéance est obligatoire.") % row_index)

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            value = value.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

        raise ValidationError(_("Ligne %s : la date d'échéance est invalide.") % row_index)

    def action_import_leasing_echeancier_excel(self):
        for rec in self:
            if rec.vehicle_purchase_type != 'leasing_contract':
                raise ValidationError(_("Cette action est réservée aux contrats leasing."))

            if rec.is_leasing_bill_order:
                raise ValidationError(_("Cette action n'est pas autorisée sur un BL issu d'une échéance."))

            if not rec.leasing_amortization_attachment:
                raise ValidationError(_("Veuillez joindre le tableau d'amortissement Excel dans l'onglet Documents leasing."))

            if not rec.vehicle_id:
                raise ValidationError(_("Veuillez sélectionner un véhicule."))

            if not rec.leasing_months or rec.leasing_months <= 0:
                raise ValidationError(_("Veuillez renseigner un nombre de mois valide."))

            if rec.leasing_amount_debit_ttc <= 0:
                raise ValidationError(_("Veuillez renseigner le montant prélèvement TTC du système."))

            try:
                file_data = base64.b64decode(rec.leasing_amortization_attachment)
                workbook = load_workbook(filename=io.BytesIO(file_data), data_only=True)
                sheet = workbook.active
            except Exception as e:
                raise ValidationError(_("Impossible de lire le fichier Excel : %s") % str(e))

            rows_to_import = []

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                sequence = row[0] if len(row) > 0 else None
                raw_date_echeance = row[1] if len(row) > 1 else None
                raw_capital = row[2] if len(row) > 2 else None
                raw_interet = row[3] if len(row) > 3 else None
                raw_montant_ttc = row[4] if len(row) > 4 else None
                raw_tva = row[5] if len(row) > 5 else 0.0

                if not any(v not in (None, '', False) for v in [
                    sequence,
                    raw_date_echeance,
                    raw_capital,
                    raw_interet,
                    raw_montant_ttc,
                    raw_tva,
                ]):
                    continue

                if sequence in (None, '', False):
                    raise ValidationError(_("Ligne %s : le numéro d'échéance est obligatoire.") % row_index)

                try:
                    sequence = int(sequence)
                except Exception:
                    raise ValidationError(_("Ligne %s : le numéro d'échéance doit être numérique.") % row_index)

                date_echeance = rec._convert_excel_date(raw_date_echeance, row_index)

                if raw_capital in (None, ''):
                    raise ValidationError(_("Ligne %s : le capital est obligatoire.") % row_index)

                if raw_interet in (None, ''):
                    raise ValidationError(_("Ligne %s : l'intérêt est obligatoire.") % row_index)

                if raw_montant_ttc in (None, ''):
                    raise ValidationError(_("Ligne %s : le montant TTC est obligatoire.") % row_index)

                try:
                    capital_amount = float(raw_capital)
                except Exception:
                    raise ValidationError(_("Ligne %s : le capital doit être numérique.") % row_index)

                try:
                    interest_amount = float(raw_interet)
                except Exception:
                    raise ValidationError(_("Ligne %s : l'intérêt doit être numérique.") % row_index)

                try:
                    amount_total = float(raw_montant_ttc)
                except Exception:
                    raise ValidationError(_("Ligne %s : le montant TTC doit être numérique.") % row_index)

                try:
                    tva_import_amount = float(raw_tva or 0.0)
                except Exception:
                    raise ValidationError(_("Ligne %s : la TVA doit être numérique.") % row_index)

                if capital_amount < 0:
                    raise ValidationError(_("Ligne %s : le capital ne peut pas être négatif.") % row_index)

                if interest_amount < 0:
                    raise ValidationError(_("Ligne %s : l'intérêt ne peut pas être négatif.") % row_index)

                if amount_total < 0:
                    raise ValidationError(_("Ligne %s : le montant TTC ne peut pas être négatif.") % row_index)

                if tva_import_amount < 0:
                    raise ValidationError(_("Ligne %s : la TVA ne peut pas être négative.") % row_index)

                if tva_import_amount > amount_total:
                    raise ValidationError(_("Ligne %s : la TVA ne peut pas dépasser le montant TTC.") % row_index)

                if round(capital_amount + interest_amount, 2) != round(amount_total, 2):
                    raise ValidationError(_("Ligne %s : Capital + Intérêt doit être égal au Montant TTC.") % row_index)

                if round(amount_total, 2) != round(rec.leasing_amount_debit_ttc, 2):
                    raise ValidationError(_(
                        "Ligne %s : le montant TTC (%s) doit être égal au montant prélèvement système (%s)."
                    ) % (row_index, amount_total, rec.leasing_amount_debit_ttc))

                rows_to_import.append({
                    'sequence': sequence,
                    'date_echeance': date_echeance,
                    'capital_amount': capital_amount,
                    'interest_amount': interest_amount,
                    'amount_total': amount_total,
                    'tva_import_amount': tva_import_amount,
                    'amount_penalite': 0.0,
                })

            if not rows_to_import:
                raise ValidationError(_("Aucune ligne valide n'a été trouvée dans le fichier Excel."))

            if len(rows_to_import) != int(rec.leasing_months):
                raise ValidationError(_(
                    "Le nombre de lignes de l'Excel (%s) doit être égal au nombre de mois du contrat (%s)."
                ) % (len(rows_to_import), int(rec.leasing_months)))

            sequences = [r['sequence'] for r in rows_to_import]
            if len(sequences) != len(set(sequences)):
                raise ValidationError(_("Le fichier Excel contient des numéros d'échéance en doublon."))

            rows_to_import = sorted(rows_to_import, key=lambda x: x['sequence'])

            if not rec.leasing_first_debit_date:
                raise ValidationError(_("Veuillez renseigner la date du 1er prélèvement sur le contrat leasing."))

            first_excel_date = rows_to_import[0]['date_echeance']
            if first_excel_date != rec.leasing_first_debit_date:
                raise ValidationError(_(
                    "La première date d'échéance du tableau d'amortissement (%s) "
                    "doit être identique à la date du 1er prélèvement du contrat (%s)."
                ) % (
                    first_excel_date.strftime('%d/%m/%Y'),
                    rec.leasing_first_debit_date.strftime('%d/%m/%Y')
                ))

            if rec.leasing_echeance_ids.filtered(lambda e: e.state == 'paid' or e.amount_paid > 0):
                raise ValidationError(_("Impossible de réimporter : certaines échéances sont déjà payées."))

            rec.leasing_echeance_ids.unlink()

            for line in rows_to_import:
                rec.env['leasing.echeance'].create({
                    'purchase_order_id': rec.id,
                    'sequence': line['sequence'],
                    'date_echeance': line['date_echeance'],
                    'date_prelevement_reel': line['date_echeance'],
                    'capital_amount': line['capital_amount'],
                    'interest_amount': line['interest_amount'],
                    'amount_total': line['amount_total'],
                    'tva_import_amount': line['tva_import_amount'],
                    'amount_penalite': 0.0,
                    'tax_ids': [(6, 0, rec.leasing_tax_ids.ids)],
                    'payment_mode': 'manuel',
                    'note': _("Importé depuis le tableau d'amortissement Excel"),
                })

    @api.onchange('leasing_amount_debit_ttc', 'leasing_tax_ids')
    def _onchange_leasing_debit_amounts(self):
        for rec in self:
            ttc = rec.leasing_amount_debit_ttc or 0.0
            total_tax_percent = sum(rec.leasing_tax_ids.mapped('amount'))

            if total_tax_percent:
                ht = ttc / (1 + (total_tax_percent / 100.0))
                tax_amount = ttc - ht
            else:
                ht = ttc
                tax_amount = 0.0

            rec.leasing_amount_debit_ht = ht
            rec.leasing_debit_tax_amount = tax_amount

    @api.constrains('vehicle_purchase_type', 'leasing_amount_debit_ttc', 'leasing_tax_ids')
    def _check_leasing_debit_values(self):
        for rec in self:
            if rec.vehicle_purchase_type == 'leasing_contract':
                if rec.leasing_amount_debit_ttc < 0:
                    raise ValidationError(_("Le montant prélèvement TTC ne peut pas être négatif."))

    @api.depends('leasing_allowed_km', 'leasing_actual_km')
    def _compute_km_info(self):
        for rec in self:
            extra = (rec.leasing_actual_km or 0.0) - (rec.leasing_allowed_km or 0.0)
            rec.leasing_extra_km = extra if extra > 0 else 0.0

    @api.depends(
        'leasing_echeance_ids.amount_total',
        'leasing_echeance_ids.amount_paid',
        'leasing_echeance_ids.amount_due',
        'leasing_echeance_ids.state',
    )
    def _compute_leasing_summary(self):
        for rec in self:
            echeances = rec.leasing_echeance_ids

            rec.leasing_echeance_count = len(echeances)
            rec.leasing_paid_echeance_count = len(echeances.filtered(lambda e: e.state == 'paid'))
            rec.leasing_remaining_echeance_count = len(echeances.filtered(lambda e: e.state != 'paid'))
            rec.leasing_late_echeance_count = len(echeances.filtered(lambda e: e.state == 'late'))

            rec.leasing_total_amount = sum(echeances.mapped('amount_total'))
            rec.leasing_total_paid = sum(echeances.mapped('amount_paid'))
            rec.leasing_total_remaining = sum(echeances.mapped('amount_due'))

    def action_generate_leasing_echeancier(self):
        for rec in self:
            if rec.vehicle_purchase_type != 'leasing_contract':
                raise ValidationError(_("Cette action est réservée aux contrats leasing."))

            if not rec.leasing_amortization_attachment:
                raise ValidationError(_("Veuillez importer le tableau d'amortissement dans l'onglet Documents leasing."))

            rec.action_import_leasing_echeancier_excel()

    def _prepare_leasing_penalty_order_line_vals(self, echeance):
        self.ensure_one()

        if not self.vehicle_id:
            raise ValidationError(_("Aucun véhicule n'est lié au contrat leasing."))

        product = self.env['product.product'].search([
            ('purchase_ok', '=', True),
            ('name', '=', 'Pénalité leasing'),
        ], limit=1)

        if not product:
            raise ValidationError(_("Créez d'abord un produit achat de type service nommé 'Pénalité leasing'."))

        if not product.uom_id:
            raise ValidationError(_("Le produit 'Pénalité leasing' n'a pas d'unité de mesure."))

        vehicle_name = (
            self.vehicle_id.display_name
            or self.vehicle_id.license_plate
            or self.vehicle_id.name
            or ''
        )

        return {
            'product_id': product.id,
            'name': _("Pénalité leasing - %(veh)s - Echéance %(seq)s - %(date)s") % {
                'veh': vehicle_name,
                'seq': echeance.sequence,
                'date': echeance.date_echeance or '',
            },
            'product_qty': 1.0,
            'product_uom_id': product.uom_id.id,
            'price_unit': echeance.amount_penalite or 0.0,
            'tax_ids': [(6, 0, [])],
            'date_planned': fields.Datetime.now(),
        }

    def button_confirm(self):
        return super().button_confirm()

    def action_view_leasing_echeances(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Echéancier leasing'),
            'res_model': 'leasing.echeance',
            'view_mode': 'list,form',
            'domain': [('purchase_order_id', '=', self.id)],
            'context': {
                'default_purchase_order_id': self.id,
            },
            'target': 'current',
        }

    def action_view_leasing_payments(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Paiements leasing'),
            'res_model': 'leasing.payment',
            'view_mode': 'list,form',
            'domain': [('purchase_order_id', '=', self.id)],
            'target': 'current',
        }